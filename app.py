from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from openpyxl.chart import BarChart, Reference
from openpyxl.chart import PieChart
import csv

planilha = Workbook()

aba = planilha.active

aba.title = "Relatório de Vendas"

aba["A18"] = f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
cabecalhos = ["Produto", "Preço", "Quantidade", "Status", "Total"]
for coluna, titulo in enumerate(cabecalhos, start=1):
    celula = aba.cell(row=1, column=coluna)

    celula.value = titulo

    celula.font = Font(
        bold=True,
        color="FFFFFF"
    )

    celula.fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    celula.alignment = Alignment(
        horizontal="center"
    )

aba.column_dimensions["A"].width = 35
aba.column_dimensions["B"].width = 15
aba.column_dimensions["C"].width = 15
aba.column_dimensions["D"].width = 20
aba.column_dimensions["E"].width = 15


produtos = [
    ["Mouse Gamer", 120, 15],
    ["Teclado Mecânico", 350, 10],
    ["Monitor 24", 900, 8],
    ["Monitor 27", 1400, 5],
    ["Notebook Dell", 4200, 3],
    ["Notebook Lenovo", 3800, 4],
    ["Headset Gamer", 280, 12],
    ["Webcam Full HD", 180, 20],
    ["SSD 1TB", 450, 18],
    ["Memória RAM 16GB", 320, 25],
    ["Placa de Vídeo RTX 4060", 2500, 2],
    ["Cadeira Gamer", 950, 6]
]

produtos.sort(key=lambda produto: produto[1] * produto[2], reverse=True)

borda = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

linha = 2

for produto in produtos:

    nome = produto[0]
    preco = produto[1]
    quantidade = produto[2]

    total = preco * quantidade

    if quantidade < 5:
        status = "CRITICO"

    elif quantidade < 10:
        status = "ATENCAO"

    else:
        status = "OK"

    aba[f"A{linha}"] = nome
    aba[f"B{linha}"] = preco
    aba[f"C{linha}"] = quantidade
    aba[f"D{linha}"] = status
    aba[f"E{linha}"] = total

    if status == "CRITICO":

        aba[f"D{linha}"].fill = PatternFill(
            start_color="F4CCCC",
            end_color="F4CCCC",
            fill_type="solid"
        )

    elif status == "ATENCAO":

        aba[f"D{linha}"].fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid"
        )

    else:

        aba[f"E{linha}"].fill = PatternFill(
            start_color="D9EAD3",
            end_color="D9EAD3",
            fill_type="solid"
        )

    aba[f"B{linha}"].number_format = 'R$ #,##0.00'
    aba[f"E{linha}"].number_format = 'R$ #,##0.00'

    for coluna in ["A", "B", "C", "D", "E"]:
        aba[f"{coluna}{linha}"].alignment = Alignment(horizontal="center")
        aba[f"{coluna}{linha}"].border = borda

    linha += 1

aba[f"C{linha}"] = "Total Geral"

aba[f"E{linha}"] = f"=SUM(E2:E{linha-1})"

aba[f"E{linha}"].font = Font(
    bold=True,
    color="FFFFFF"
)

aba[f"D{linha}"].font = Font(
    bold=True,
    color="FFFFFF"
)

aba[f"E{linha}"] = f"=SUM(E2:E{linha-1})"

aba[f"E{linha}"].number_format = 'R$ #,##0.00'

aba[f"E{linha}"].fill = PatternFill(
    start_color="D9EAD3",
    end_color="D9EAD3",
    fill_type="solid"
)

aba[f"E{linha}"].fill = PatternFill(
    start_color="D9EAD3",
    end_color="D9EAD3",
    fill_type="solid"
)

aba[f"C{linha}"].alignment = Alignment(horizontal="center")
aba[f"E{linha}"].alignment = Alignment(horizontal="center")

aba[f"C{linha}"].border = borda
aba[f"E{linha}"].border = borda

aba.freeze_panes = "A2"

aba.auto_filter.ref = f"A1:E{linha-1}"

aba.sheet_view.showGridLines = False

aba.sheet_view.zoomScale = 120

grafico = BarChart()

grafico.title = "Vendas por Produto"
grafico.y_axis.title = "Valor Total"
grafico.x_axis.title = "Produtos"

dados = Reference(
    aba,
    min_col=5,
    min_row=2,
    max_row=linha-1
)

categorias = Reference(
    aba,
    min_col=1,
    min_row=2,
    max_row=linha-1
)

grafico.add_data(dados, titles_from_data=False)
grafico.set_categories(categorias)

aba.add_chart(grafico, "F2")

grafico_pizza = PieChart()

grafico_pizza.title = "Participação nas Vendas"

dados_pizza = Reference(
    aba,
    min_col=4,
    min_row=2,
    max_row=linha-1
)

categorias_pizza = Reference(
    aba,
    min_col=1,
    min_row=2,
    max_row=linha-1
)

grafico_pizza.add_data(dados_pizza)
grafico_pizza.set_categories(categorias_pizza)

aba.add_chart(grafico_pizza, "F20")

with open("vendas.csv", "w", newline="", encoding="utf-8") as arquivo:

    escritor = csv.writer(arquivo)

    escritor.writerow(["Produto", "Preço", "Quantidade", "Total"])

    for produto in produtos:

        nome = produto[0]
        preco = produto[1]
        quantidade = produto[2]

        total = preco * quantidade

        escritor.writerow(
            [nome, preco, quantidade, total]
        )

resumo = planilha.create_sheet("Resumo")

resumo["A1"] = "RESUMO EXECUTIVO"

resumo["A1"].font = Font(
    bold=True,
    size=16,
    color="FFFFFF"
)

resumo["A1"].fill = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

resumo["A3"] = "Quantidade de Produtos"
resumo["B3"] = len(produtos)

faturamento_total = sum (
    produto[1] * produto[2]
    for produto in produtos
)

resumo["A4"] = "Faturamento Total"
resumo["B4"] = faturamento_total
resumo["B4"].number_format = 'R$ #,##0.00'

produto_top = max(
    produtos,
    key=lambda produto: produto[1] * produto[2]
)

resumo["A5"] = "Produto Destaque"
resumo["B5"] = produto_top[0]

quantidade_total = sum(
    produto[2]
    for produto in produtos
)

resumo["A6"] = "Quantidade Vendida"
resumo["B6"] = quantidade_total

resumo["A8"] = "TOP 3 PRODUTOS"
resumo["A8"].font = Font(bold=True)

top3 = sorted(
    produtos,
    key=lambda produto: produto[1] * produto[2],
    reverse=True
)[:3]

for i, produto in enumerate(top3, start=9):

    resumo[f"A{i}"] = f"{i-8}. {produto[0]}"

for celula in ["A3", "A4", "A5", "A6"]:
    resumo[celula].font = Font(bold=True)

with open("relatorio.txt", "w", encoding="utf-8") as arquivo:

    arquivo.write("RELATÓRIO DE VENDAS\n\n")

    faturamento_total = 0

    for produto in produtos:
        
        nome = produto[0]
        preco = produto[1]
        quantidade = produto[2]

        total = preco * quantidade

        faturamento_total += total

        arquivo.write(
            f"Produto: {nome} | Quantidade: {quantidade} | Total: R${total}\n"
        )

    arquivo.write(f"\nFaturamento Total: R${faturamento_total}")

ranking = planilha.create_sheet("Ranking")

ranking["A1"] = "Produto"
ranking["B1"] = "Faturamento"

ranking["A1"].font = Font(bold=True)
ranking["B1"].font = Font(bold=True)

produtos_ordenados = sorted(
    produtos,
    key=lambda produto: produto[1] * produto[2],
    reverse=True
)

linha_ranking = 2

for produto in produtos_ordenados:

    faturamento = produto[1] * produto[2]

    ranking[f"A{linha_ranking}"] = produto[0]
    ranking[f"B{linha_ranking}"] = faturamento

    ranking[f"B{linha_ranking}"].number_format = 'R$ #,##0.00'

    linha_ranking += 1

ranking["C2"] = "🥇"
ranking["C3"] = "🥈"
ranking["C4"] = "🥉"


planilha.save("vendas.xlsx")

print("Planilha criada com sucesso!")